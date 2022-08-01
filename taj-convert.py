import openpyexcel as ope, random


from openpyxl import workbook


wb = ope.load_workbook('Addresses.xlsx')
sheet = wb['sheet1']
cell = sheet['a1']
cell = sheet.cell(1 , 1)
mask = '255.255.255.255'
for row in range(2, 14):
    cell = sheet.cell(row, 4)
    cell.value = mask
    print(cell.value)
    
wb.save('Addresses2.txt')
   
    

#for number in range(1,10 + 1):
 #   print(number)
 
"""
number = 1
while number < 11:
    print(number)
    number += 1
    """

Winner = ["Dayo", "Segun", "Tobi", "Lekan", "Abraham"]
print(random.choice(Winner))

